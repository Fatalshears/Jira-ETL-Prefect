###########################################
#Author: Nguyen Anh Duc (MS/EDA32-XC)
###########################################
import typing
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime
import os
from docxtpl import DocxTemplate


# since data may contains of empty list or None datatype, remove these data with empty string, 
# convert non-empty list into string so data can be written using openpyxl
def data_check(data):
    refined_data = []
    for x in data:
        if x == None or x == []:
            x = ''
        elif isinstance(x, typing.List) and len(x) > 0:
            x = ', '.join([ele for ele in x])

        refined_data.append(x)
    return refined_data

def update_test_summary(data, sys_bug_count, sw_bug_count, proj_name, series, release, data_misc):
    reference_template = r'C:\Users\udn1hc\CodeProject\Jira_Reporting_Tool\Defect_Metrics_Tool\Template_Test_Summary_Report.xlsx'
    output_folder = r'\\bosch.com\dfsRB\DfsVN\LOC\Hc\RBVH\20_EDA\10_EDA1\01_Internal\98_Database\02_Test_Coordinator\Problem_Ticket_Metrics\Data'
    # output_folder = r'C:\Users\udn1hc\CodeProject\Jira_Reporting_Tool\Defect_Metrics_Tool\Data'

    test_summary_wb = load_workbook(reference_template)
    problem_list_ws = test_summary_wb['T&R_Problem_List']
    # define table headers
    problem_list_ws.append(["Issue Type", "Key", "Summary", "Assignee", "Reporter", "Status", "Resolution", "Fix Version/s", "Labels", "Component/s", "Scope", "Severity", "Safety Relevance", "Affects Version/s", "Origin", "Problem Type"])
    # table data
    for row in data:
        problem_list_ws.append(data_check(row))
    # get a timestamp for table display name, each new table should have a unique name
    current_time = datetime.now().strftime("%d%b%Y%I%M%S%p")
    displayName = "Table{0}".format(current_time)
    tab = Table(displayName=displayName, ref="A1:" + get_column_letter(len(data[0])) + str(len(data)+1))


    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleLight13", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style

    problem_list_ws.add_table(tab)

    # modify column width of table
    for col in range(1, len(data[0])+1):
        table_headers = problem_list_ws.column_dimensions[get_column_letter(col)]
        if col == 3 or col == 9 or col == 14:
            table_headers.width = 40
        elif col == 2:
            table_headers.width = 20
        else:
            table_headers.width = 15

    
    # applied color for type of issues
    system_defect_color = "66CCFF"
    software_defect_color = "00FF99"
    performance_issue_color = "FFCC00"
    system_defect_labels = ['DetectionPhase_SystemTest(SYS_TST)', 'DetectionPhase_SystemIntegrationandIntegrationTest(SYS_INT)', 'DetectionPhase_SystemApplication(SYS_APP)','InjectionPhase_Errata']
    software_defect_labels = ['DetectionPhase_SWUnitVerification(SW_UVE)', 'SW_UVE_DEFECT', 'DetectionPhase_SWIntegrationandIntegrationTest(SW_INT)', 'DetectionPhase_SWTest(SW_TST)']
    for i in range(2, len(data)+2):
        # check labels and problem type
        ticket_summary_cell = problem_list_ws.cell(row = i, column = 3)
        problem_type_cell = problem_list_ws.cell(row = i, column = 16)
        if problem_type_cell.value == 'Defect':
            if data[i-2][8] != None and any(x in system_defect_labels for x in data[i-2][8]):
                ticket_summary_cell.fill = PatternFill(start_color=system_defect_color, end_color=system_defect_color, fill_type="solid")
            elif data[i-2][8] != None and any(x in software_defect_labels for x in data[i-2][8]):
                ticket_summary_cell.fill = PatternFill(start_color=software_defect_color, end_color=software_defect_color, fill_type="solid")
        else:
            ticket_summary_cell.fill = PatternFill(start_color=performance_issue_color, end_color=performance_issue_color, fill_type="solid")

    # create table legend
    legend_system_defect_color = problem_list_ws.cell(row = len(data)+4, column = 3)
    legend_system_defect_color.value = "Defect in System Test"
    legend_system_defect_color.fill = PatternFill(start_color=system_defect_color, end_color=system_defect_color, fill_type="solid")

    legend_software_defect_color = problem_list_ws.cell(row = len(data)+5, column = 3)
    legend_software_defect_color.value = "Defect in SW Test"
    legend_software_defect_color.fill = PatternFill(start_color=software_defect_color, end_color=software_defect_color, fill_type="solid")

    legend_erformance_issue_color = problem_list_ws.cell(row = len(data)+6, column = 3)
    legend_erformance_issue_color.value = "Performance Issue, Other"
    legend_erformance_issue_color.fill = PatternFill(start_color=performance_issue_color, end_color=performance_issue_color, fill_type="solid")

    # update T&R_Defect_Status sheet
    defect_status_ws = test_summary_wb['T&R_Defect_Status']
    defect_status_ws['A3'].value = release
    defect_status_ws['B3'].value = data_misc[0]
    defect_status_ws['C3'].value = data_misc[1]
    defect_status_ws['D3'].value = data_misc[2]
    defect_status_ws['E3'].value = data_misc[3]
    defect_status_ws['F3'].value = sys_bug_count

    defect_status_ws['H3'].value = release
    defect_status_ws['I3'].value = data_misc[4]
    defect_status_ws['J3'].value = data_misc[5]
    defect_status_ws['K3'].value = data_misc[6]
    defect_status_ws['L3'].value = data_misc[7]
    defect_status_ws['M3'].value = sw_bug_count

    # get timestamp and add it to file name, each new file should have a unique name
    current_time = datetime.now().strftime("[%d-%b-%Y_%I-%M-%S-%p]")
    release_folder_ouput = output_folder + f'\\{proj_name}' + f'\\{series}' + f'\\{release}'
    if not os.path.isdir(release_folder_ouput):
        os.makedirs(release_folder_ouput)
    report_file_name = r'{0}\{1}_{2}_{3}_Test_Summary_Report.xlsx'.format(release_folder_ouput,current_time,series,release)
    test_summary_wb.save(report_file_name)

def update_release_note(data, bug_strong, bug_medium, bug_minor, perf_strong, perf_medium, perf_minor, proj_name, series, release, data_misc):
    reference_template = r'C:\Users\udn1hc\CodeProject\Jira_Reporting_Tool\Defect_Metrics_Tool\Template_SW_Release_Note.docx'
    output_folder = r'\\bosch.com\dfsRB\DfsVN\LOC\Hc\RBVH\20_EDA\10_EDA1\01_Internal\98_Database\02_Test_Coordinator\Problem_Ticket_Metrics\Data'
    # output_folder = r'C:\Users\udn1hc\CodeProject\Jira_Reporting_Tool\Defect_Metrics_Tool\Data'

    doc = DocxTemplate(reference_template)

    internal_lim_rows = []
    count = 0
    for item in data:
        count = count+1
        internal_lim_rows.append({"count": count,"key": item[0], "summary": item[1], "severity": item[2], "safety": item[3]})

    context = {
        "defect_stg_cl": bug_strong,
        "defect_med_cl": bug_medium,
        "defect_mnr_cl": bug_minor,
        "defect_tt_cl": bug_strong+bug_medium+bug_minor,
        "defect_stg_o": data_misc[8],
        "defect_med_o": data_misc[9],
        "defect_mnr_o": data_misc[10],
        "defect_tt_o": data_misc[8]+data_misc[9]+data_misc[10],
        "defect_stg_ip": data_misc[11],
        "defect_med_ip": data_misc[12],
        "defect_mnr_ip": data_misc[13],
        "defect_tt_ip": data_misc[11]+data_misc[12]+data_misc[13],
        "perf_stg_cl": perf_strong,
        "perf_med_cl": perf_medium,
        "perf_mnr_cl": perf_minor,
        "perf_tt_cl": perf_strong+perf_medium+perf_minor,
        "perf_stg_o": data_misc[14],
        "perf_med_o": data_misc[15],
        "perf_mnr_o": data_misc[16],
        "perf_tt_o": data_misc[14]+data_misc[15]+data_misc[16],
        "perf_stg_ip": data_misc[17],
        "perf_med_ip": data_misc[18],
        "perf_mnr_ip": data_misc[19],
        "perf_tt_ip": data_misc[17]+data_misc[18]+data_misc[19],
        "internal_lim_tbl_rows": internal_lim_rows
    }

    doc.render(context)
    # get timestamp and add it to file name, each new file should have a unique name
    current_time = datetime.now().strftime("[%d-%b-%Y_%I-%M-%S-%p]")
    release_folder_ouput = output_folder + f'\\{proj_name}' + f'\\{series}' + f'\\{release}'
    if not os.path.isdir(release_folder_ouput):
        os.makedirs(release_folder_ouput)
    report_file_name = r'{0}\{1}_{2}_{3}_SW_Release_Note.docx'.format(release_folder_ouput,current_time,series,release)
    doc.save(report_file_name)


# data = [('Problem', 'FRGVCAGA-397', 'QA_Level7-9_Warnings: SIT-RPM-DSM: [CA_B316MCA1_FR_BL03_V01]', 'WUT6SZH', 'DAN9HC', 'Open', 'Unresolved', [], ['DetectionPhase_SWUnitVerification(SW_UVE)', 'Static_Analysis'], [], None, 'Medium', 'No', ['B316_BL02_V1', 'CA_B316MCA1_BL01_V02', 'B316_BL03_V1'], 'Project Internal', 'Defect')]
# modify_test_summary(data, 4, 10, 'BL09')